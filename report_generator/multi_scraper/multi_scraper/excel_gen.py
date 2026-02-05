import json
import os
import re
from openpyxl import Workbook
from openai import OpenAI
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
# ================= GPT CLIENT =================
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# ================= PATHS =================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_FOLDER = os.path.join(BASE_DIR, "..", "dominating_region")


# ================= HELPERS =================
def get_unit_suffix(unit):
    if not unit:
        return ""
    unit = unit.lower()
    if unit == "million":
        return "Mn"
    elif unit == "billion":
        return "Bn"
    elif unit == "trillion":
        return "Tn"
    return ""


def clean_filename(name):
    return re.sub(r'[\\/:*?"<>|]', '', name).strip()


def generate_template_excel(market_name, unit="Million"):
    """Generate a basic template Excel when no dominating_region data is available"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Add header
    ws['A1'] = f"{market_name} - Regional Analysis Template"
    ws['A2'] = "Note: This is a template file. No regional data available."
    ws['A4'] = "Region"
    ws['B4'] = f"Market Size ({unit})"
    
    # Add sample regions
    regions = ["North America", "Europe", "Asia Pacific", "Latin America", "Middle East & Africa"]
    for idx, region in enumerate(regions, start=5):
        ws[f'A{idx}'] = region
        ws[f'B{idx}'] = "N/A"
    
    return wb


# ================= GPT SINGLE CALL (DOM + FAST) =================
def get_gpt_dominating_and_fastest_sub_segments(market_name, sub_segments):
    if not sub_segments:
        return [], []

    prompt = f"""
    Act like a research analyst.

    From the following sub segments of the {market_name}, do TWO things:

    1. Rank sub segments based on DOMINANCE
    2. Rank sub segments based on FASTEST GROWTH

    Sub segments:
    {', '.join(sub_segments)}

    Output format STRICTLY like this:

    Dominating:
    segment1, segment2, segment3

    Fastest Growing:
    segment1, segment2, segment3

    Rules:
    - Output ONLY sub segment names
    - No explanations
    """

    response = client.chat.completions.create(
        model="gpt-5-mini",
        messages=[{"role": "user", "content": prompt}],
    )

    text = response.choices[0].message.content.strip()

    dom_list, fast_list = [], []

    dom_match = re.search(r"Dominating:\s*(.*?)(Fastest Growing:|$)", text, re.S)
    fast_match = re.search(r"Fastest Growing:\s*(.*)", text, re.S)

    if dom_match:
        dom_raw = dom_match.group(1)
        dom_list = [
            s.strip()
            for s in dom_raw.replace("\n", ",").split(",")
            if s.strip() in sub_segments
        ]

    if fast_match:
        fast_raw = fast_match.group(1)
        fast_list = [
            s.strip()
            for s in fast_raw.replace("\n", ",").split(",")
            if s.strip() in sub_segments
        ]

    # Safe fallback
    if not dom_list:
        dom_list = sub_segments
    if not fast_list:
        fast_list = sub_segments

    return dom_list, fast_list


# ================= MAIN EXCEL GENERATOR =================
def generate_excel(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    market_name = data.get("market_name", "")
    unit = data.get("unit", "")
    regional_ranking = data.get("REGIONAL_RANKING", {})
    europe_classification = data.get("EUROPE_COUNTRY_CLASSIFICATION", {})
    segments = data.get("SEGMENTS", {})

    first_region = regional_ranking.get("first", "")
    second_region = regional_ranking.get("second", "")

    first_segment = next(iter(segments.keys()), "")
    raw_sub_segments = segments.get(first_segment, [])

    # ===== SINGLE GPT CALL =====
    dominating_segments, fastest_growing_segments = (
        get_gpt_dominating_and_fastest_sub_segments(
            market_name, raw_sub_segments
        )
    )
    
    # Print dominating and fastest growing to terminal
    print("\n" + "="*60)
    print(f"MARKET: {market_name}")
    print("="*60)
    print("\nDOMINATING SEGMENTS:")
    for i, seg in enumerate(dominating_segments, 1):
        print(f"  {i}. {seg}")
    print("\nFASTEST GROWING SEGMENTS:")
    for i, seg in enumerate(fastest_growing_segments, 1):
        print(f"  {i}. {seg}")
    print("="*60 + "\n")

    unit_suffix = get_unit_suffix(unit)
    years = list(range(2025, 2034))

    wb = Workbook()

    # ================= SHEET 1 =================
    ws1 = wb.active
    ws1.title = "global_market_by_region"

    ws1["A1"] = "position"
    ws1["B1"] = 1
    ws1["A2"] = "above_name"
    ws1["B2"] = f"{market_name} ($ {unit_suffix})"
    ws1["A3"] = "below_name"

    ws1["A5"] = "Year"
    ws1["B5"] = regional_ranking.get("first", "")
    ws1["C5"] = regional_ranking.get("second", "")
    ws1["D5"] = regional_ranking.get("third", "")
    ws1["E5"] = "Middle East & Africa"
    ws1["F5"] = "Latin America"

    for i, year in enumerate(years):
        r = 6 + i
        ws1[f"A{r}"] = year
        ws1[f"B{r}"] = 11 + i
        ws1[f"C{r}"] = 10 + i
        ws1[f"D{r}"] = 9 + i
        ws1[f"E{r}"] = 8 + i
        ws1[f"F{r}"] = 7 + i

    # ================= SHEET 2 =================
    ws2 = wb.create_sheet("country_share")

    ws2["A1"] = "position"
    ws2["B1"] = 2
    ws2["A2"] = "above_name"
    ws2["B2"] = f"Country Share for {first_region} Region (%)"
    ws2["A3"] = "below_name"
    ws2["A5"] = "Country"
    ws2["B5"] = 2025

    if first_region == "North America":
        ws2["A6"], ws2["A7"] = "US", "Canada"
        ws2["B6"], ws2["B7"] = 25, 15
    elif first_region == "Asia Pacific":
        ws2["A6"], ws2["A7"] = "Japan", "South Korea"
        ws2["B6"], ws2["B7"] = 25, 15
    elif first_region == "Europe":
        dom = next((k for k, v in europe_classification.items() if v == "dominant"), "")
        fast = next((k for k, v in europe_classification.items() if v == "fastest_growing"), "")
        emer = next((k for k, v in europe_classification.items() if v == "emerging"), "")
        ws2["A6"], ws2["A7"], ws2["A8"] = dom, fast, emer
        ws2["B6"], ws2["B7"], ws2["B8"] = 25, 15, 5

    # ================= SHEET 3 (DOMINATING) =================
    ws3 = wb.create_sheet("Segment_1_Share")

    ws3["A1"] = "position"
    ws3["B1"] = 3
    ws3["A2"] = "above_name"
    ws3["B2"] = f"{market_name} By {first_segment} ($ {unit_suffix})"
    ws3["A3"] = "below_name"
    ws3["A5"] = "Year"

    for i, seg in enumerate(dominating_segments[:5]):
        ws3.cell(row=5, column=2 + i, value=seg)

    for i, year in enumerate(years):
        r = 6 + i
        ws3[f"A{r}"] = year
        for j in range(len(dominating_segments[:5])):
            ws3.cell(row=r, column=2 + j, value=12 - j + i)

    # ================= SHEET 4 (FASTEST GROWING) =================
    ws4 = wb.create_sheet("cagr")

    ws4["A1"] = "position"
    ws4["B1"] = 4
    ws4["A2"] = "above_name"
    ws4["B2"] = f"{market_name} By {first_segment} (%)"
    ws4["A3"] = "below_name"
    ws4["A5"] = "Year"

    for i, seg in enumerate(fastest_growing_segments[:5]):
        ws4.cell(row=5, column=2 + i, value=seg)

    for i, year in enumerate(years):
        r = 6 + i
        ws4[f"A{r}"] = year
        for j in range(len(fastest_growing_segments[:5])):
            ws4.cell(row=r, column=2 + j, value=12 - j + i)

    # ================= SHEET 5 (DOMINATING) =================
    ws5 = wb.create_sheet("Segment_2_Share")

    ws5["A1"] = "position"
    ws5["B1"] = 5
    ws5["A2"] = "above_name"
    ws5["B2"] = f"{market_name} By {first_segment}"
    ws5["A3"] = "below_name"

    ws5["A5"] = "Size"
    ws5["B5"] = "Segment Name"

    sizes = [1000, 900, 800, 700, 600]
    for i, seg in enumerate(dominating_segments[:5]):
        ws5[f"A{6 + i}"] = sizes[i]
        ws5[f"B{6 + i}"] = seg

    # ================= SHEET 6 =================
    ws6 = wb.create_sheet("worldmap")

    ws6["A1"] = "position"
    ws6["B1"] = 6
    ws6["A2"] = "above_name"
    ws6["B2"] = f"{market_name} By Geography"
    ws6["A3"] = "below_name"

    ws6["A5"] = "Size"
    ws6["B5"] = "Region"

    ws6["A6"], ws6["A7"], ws6["A8"], ws6["A9"] = 900, 900, 700, 700

    def europe_top_two():
        dom = next((k for k, v in europe_classification.items() if v == "dominant"), "")
        fast = next((k for k, v in europe_classification.items() if v == "fastest_growing"), "")
        return dom, fast

    if first_region == "North America":
        ws6["B6"], ws6["B7"] = "United States", "Canada"
    elif first_region == "Europe":
        ws6["B6"], ws6["B7"] = europe_top_two()
    elif first_region == "Asia Pacific":
        ws6["B6"], ws6["B7"] = "Japan", "South Korea"

    if second_region == "North America":
        ws6["B8"], ws6["B9"] = "United States", "Canada"
    elif second_region == "Europe":
        ws6["B8"], ws6["B9"] = europe_top_two()
    elif second_region == "Asia Pacific":
        ws6["B8"], ws6["B9"] = "Japan", "South Korea"

    return wb, market_name


# ================= ENTRY POINT =================
def main():
    json_files = [f for f in os.listdir(JSON_FOLDER) if f.endswith(".json")]
    if not json_files:
        print("No JSON files found.")
        return

    json_path = os.path.join(JSON_FOLDER, json_files[0])
    wb, market_name = generate_excel(json_path)

    output_file = f"{clean_filename(market_name)}.xlsx"
    wb.save(output_file)
    print(f"Excel file generated successfully: {output_file}")


if __name__ == "__main__":
    main()
